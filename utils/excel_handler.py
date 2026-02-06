import pandas as pd
from typing import Tuple, Optional
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, PatternFill
from copy import copy
from .dados_estaticos import COLUNAS_EXCEL
from .calculos import parse_data_excel


class ExcelHandler:
    """Classe para manipular a planilha Excel de mortes no trânsito."""

    def __init__(self, caminho_arquivo: str = None):
        """
        Inicializa o handler do Excel.

        Args:
            caminho_arquivo: Caminho do arquivo Excel (opcional)
        """
        self.caminho_arquivo = caminho_arquivo
        self.df = None
        self.dados_carregados = False

    def carregar_arquivo(self, caminho: str) -> Tuple[bool, str]:
        """
        Carrega o arquivo Excel e valida sua estrutura.

        Args:
            caminho: Caminho do arquivo Excel

        Returns:
            Tupla (sucesso, mensagem)
        """
        try:
            # Carrega o arquivo
            self.df = pd.read_excel(caminho)
            self.caminho_arquivo = caminho

            # Valida as colunas
            if len(self.df.columns) != 33:
                return False, f"Arquivo possui {len(self.df.columns)} colunas. Esperado: 33 colunas."

            # Verifica se os nomes das colunas batem (ignorando diferenças de espaços)
            colunas_arquivo = [col.strip() for col in self.df.columns]
            colunas_esperadas = [col.strip() for col in COLUNAS_EXCEL]

            if colunas_arquivo != colunas_esperadas:
                return False, "Estrutura de colunas do arquivo não corresponde ao esperado."

            self.dados_carregados = True
            total_registros = len(self.df)

            return True, f"Arquivo carregado com sucesso! {total_registros} registros encontrados."

        except FileNotFoundError:
            return False, "Arquivo não encontrado."
        except PermissionError:
            return False, "Sem permissão para ler o arquivo. Verifique se ele não está aberto em outro programa."
        except Exception as e:
            return False, f"Erro ao carregar arquivo: {str(e)}"

    def obter_info_arquivo(self) -> dict:
        """
        Retorna informações sobre o arquivo carregado.

        Returns:
            Dicionário com informações do arquivo
        """
        if not self.dados_carregados or self.df is None:
            return {
                'total_registros': 0,
                'ultima_data': None,
                'municipios_unicos': 0
            }

        # Processa a última data
        ultima_data = None
        if 'Data do Fato' in self.df.columns:
            datas_validas = self.df['Data do Fato'].dropna()
            if len(datas_validas) > 0:
                ultima_data_valor = datas_validas.iloc[-1]
                ultima_data = parse_data_excel(ultima_data_valor)

        return {
            'total_registros': len(self.df),
            'ultima_data': ultima_data,
            'municipios_unicos': self.df['Município do Fato'].nunique()
        }

    def inserir_registro(self, dados: dict) -> Tuple[bool, str, int]:
        """
        Insere um novo registro na planilha, ordenado por Data do Fato.

        Args:
            dados: Dicionário com os dados do novo registro

        Returns:
            Tupla (sucesso, mensagem, posicao_inserida)
        """
        if not self.dados_carregados:
            return False, "Nenhum arquivo carregado.", -1

        try:
            # Cria um novo registro como DataFrame
            novo_registro = pd.DataFrame([dados])

            # Adiciona ao DataFrame existente
            self.df = pd.concat([self.df, novo_registro], ignore_index=True)

            # Ordena por Data do Fato
            self.df['Data do Fato'] = pd.to_datetime(self.df['Data do Fato'], errors='coerce')
            self.df = self.df.sort_values(by='Data do Fato', ascending=True)
            self.df = self.df.reset_index(drop=True)

            # Encontra a posição onde o registro foi inserido
            # (procura pela vítima e data do fato)
            posicao = -1
            for idx, row in self.df.iterrows():
                if (row['Vítima'] == dados.get('Vítima') and
                    row['Data do Fato'] == pd.to_datetime(dados.get('Data do Fato'))):
                    posicao = idx + 1  # +1 porque Excel começa em 1 e tem cabeçalho
                    break

            return True, "Registro inserido com sucesso!", posicao

        except Exception as e:
            return False, f"Erro ao inserir registro: {str(e)}", -1

    def salvar_arquivo(self, caminho_destino: str = None) -> Tuple[bool, str]:
        """
        Salva o DataFrame em um arquivo Excel preservando a formatação original.

        Args:
            caminho_destino: Caminho para salvar (se None, usa o arquivo original)

        Returns:
            Tupla (sucesso, mensagem)
        """
        if not self.dados_carregados:
            return False, "Nenhum dado para salvar."

        import os

        try:
            if caminho_destino is None:
                caminho_destino = self.caminho_arquivo

            # Carrega o workbook original para copiar formatação
            wb_original = load_workbook(self.caminho_arquivo)
            ws_original = wb_original.active

            # Salva temporariamente com pandas (usa extensão .xlsx para o temp)
            temp_file = caminho_destino.replace('.xlsx', '') + '_temp.xlsx'
            self.df.to_excel(temp_file, index=False, engine='openpyxl')

            # Carrega o arquivo temporário
            wb_novo = load_workbook(temp_file)
            ws_novo = wb_novo.active

            # Copia formatação das colunas
            for col_idx in range(1, 34):  # 33 colunas
                # Copia largura da coluna
                col_letter_original = ws_original.cell(1, col_idx).column_letter
                col_letter_novo = ws_novo.cell(1, col_idx).column_letter

                if col_letter_original in ws_original.column_dimensions:
                    ws_novo.column_dimensions[col_letter_novo].width = \
                        ws_original.column_dimensions[col_letter_original].width

                # Copia formatação do cabeçalho
                cell_original = ws_original.cell(1, col_idx)
                cell_novo = ws_novo.cell(1, col_idx)

                if cell_original.font:
                    cell_novo.font = copy(cell_original.font)
                if cell_original.alignment:
                    cell_novo.alignment = copy(cell_original.alignment)
                if cell_original.fill:
                    cell_novo.fill = copy(cell_original.fill)
                if cell_original.border:
                    cell_novo.border = copy(cell_original.border)

                # Aplica formatação de data para colunas de data
                nome_coluna = self.df.columns[col_idx - 1]
                if 'Data' in nome_coluna or nome_coluna == 'Hora do fato':
                    # Pega formatação de uma célula de data do original
                    formato_original = None
                    for row_idx in range(2, min(10, ws_original.max_row + 1)):
                        cell = ws_original.cell(row_idx, col_idx)
                        if cell.value and cell.number_format:
                            formato_original = cell.number_format
                            break

                    # Aplica formato em todas as células da coluna
                    if formato_original:
                        for row_idx in range(2, ws_novo.max_row + 1):
                            cell = ws_novo.cell(row_idx, col_idx)
                            cell.number_format = formato_original

            # Formata colunas de data especificamente
            self._formatar_colunas_data(ws_novo)

            # Salva o arquivo final
            wb_novo.save(caminho_destino)

            # Remove arquivo temporário
            if os.path.exists(temp_file):
                os.remove(temp_file)

            return True, f"Arquivo salvo com sucesso em: {caminho_destino}"

        except PermissionError:
            return False, "Sem permissão para salvar o arquivo. Verifique se ele não está aberto."
        except Exception as e:
            return False, f"Erro ao salvar arquivo: {str(e)}"

    def _formatar_colunas_data(self, ws):
        """
        Formata colunas de data com o formato correto.

        Args:
            ws: Worksheet do openpyxl
        """
        # Mapeia colunas que contêm datas
        colunas_data = {
            'Data do Óbito': 'DD/MM/YYYY',
            'Data de\nNascimento': 'DD/MM/YYYY',
            'Data do Fato': 'DD/MM/YYYY',
        }

        # Encontra índices das colunas
        for col_idx, col_name in enumerate(self.df.columns, start=1):
            if col_name in colunas_data:
                formato = colunas_data[col_name]
                # Aplica formato em todas as linhas
                for row_idx in range(2, ws.max_row + 1):
                    cell = ws.cell(row_idx, col_idx)
                    cell.number_format = formato

                    # Ajusta largura da coluna para datas
                    col_letter = cell.column_letter
                    if col_letter not in ws.column_dimensions or ws.column_dimensions[col_letter].width < 12:
                        ws.column_dimensions[col_letter].width = 12

    def obter_dataframe(self) -> Optional[pd.DataFrame]:
        """
        Retorna o DataFrame atual.

        Returns:
            DataFrame ou None se não houver dados
        """
        return self.df if self.dados_carregados else None

    def obter_valores_unicos(self, coluna: str) -> list:
        """
        Retorna valores únicos de uma coluna específica.

        Args:
            coluna: Nome da coluna

        Returns:
            Lista de valores únicos ordenados
        """
        if not self.dados_carregados or coluna not in self.df.columns:
            return []

        valores = self.df[coluna].dropna().unique().tolist()
        return sorted(valores)

    def obter_ultimo_registro(self) -> Optional[dict]:
        """
        Retorna o último registro inserido.

        Returns:
            Dicionário com os dados ou None
        """
        if not self.dados_carregados or self.df is None or len(self.df) == 0:
            return None

        return self.df.iloc[-1].to_dict()
